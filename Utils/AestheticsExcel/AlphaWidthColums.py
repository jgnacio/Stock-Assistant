import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def AlphaWidthColums(xlsx_file_path, file_name, save_folder_path):
    # Cargar el archivo de Excel existente
    wb = load_workbook(xlsx_file_path)
    ws = wb.active

    # Iterar sobre todas las columnas
    for col in ws.columns:
        max_length = 0
        # Obtener la letra de la columna
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        # Ajustar el ancho con un factor multiplicador
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Verificar que exista la carpeta de salida
    if not os.path.exists(save_folder_path):
        os.makedirs(save_folder_path)

    # Guardar los cambios
    wb.save(f'{save_folder_path}{file_name}')
